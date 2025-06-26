from flask import Blueprint, request, make_response, json, jsonify, send_file
from controllers.auth import login_required  # 추가
from sqlalchemy import text
from db import engine
import os
import io
import openpyxl 
from datetime import datetime

bl_update_bp = Blueprint('bl_update', __name__)

##### 권한 체크 공통 함수 #####
def check_bl_permission(em_id, bl_id):
    """건물에 대한 권한을 체크하는 함수"""
    try:
        with engine.connect() as conn:
            # 건물의 prop_id 조회
            prop_sql = text("SELECT prop_id FROM bl WHERE bl_id = :bl_id")
            prop_result = conn.execute(prop_sql, {"bl_id": bl_id}).fetchone()
            
            if not prop_result:
                return False
                
            prop_id = prop_result['prop_id']
            
            # 권한 체크
            auth_sql = text("""
                SELECT COUNT(*) as cnt
                FROM emcontrol e
                WHERE e.em_id = :em_id AND e.prop_id = :prop_id
            """)
            auth_result = conn.execute(auth_sql, {"em_id": em_id, "prop_id": prop_id}).fetchone()
            
            has_permission = auth_result and auth_result['cnt'] > 0
            print(f"🔍 권한 체크: em_id={em_id}, bl_id={bl_id}, prop_id={prop_id}, 권한={'있음' if has_permission else '없음'}")
            return has_permission
            
    except Exception as e:
        print(f"🔴 권한 체크 중 오류 발생: {str(e)}")
        return False

##### 건물 데이터 조회 #####
@bl_update_bp.route('/bl_update/get_bl_data', methods=['POST'])
@login_required  # 추가
def get_bl_data():
    request_data = request.get_json()
    bl_id = request_data.get('bl_id')
    em_id = request_data.get('em_id')
    
    print(f"🔵 건물 데이터 조회 요청: bl_id={bl_id}, em_id={em_id}")
    
    if not bl_id or not em_id:
        return jsonify({'success': False, 'message': 'bl_id와 em_id가 필요합니다.'})
    
    try:
        # 권한 체크
        has_permission = check_bl_permission(em_id, bl_id)
        
        with engine.connect() as conn:
            sql = text("""
                SELECT 
                    bl_id, prop_id, name, zip, contact_phone, address1, address2,
                    use1, contact_fax, DATE(date_bl) as date_bl, price_book_value,
                    contact_name, maskname, count_fl, count_bf, construction_type,
                    use_fl_4, use_fl_13, bl_height, bl_depth, width_front_road,
                    width_back_road, width_side_road, ph, el, el_unit, parking_type,
                    es, es_unit, parking_unit_inner, parking_unit_outdoor,
                    cooling_type, heating_type, DATE(date_buy) as date_buy,
                    DATE(date_buy_land) as date_buy_land, DATE(date_sailed) as date_sailed,
                    DATE(date_manage_start) as date_manage_start, price_pa_land,
                    class_land, regist, district, section, landlord, design,
                    builder, elev_type, elec_type, generator_type, roof_type,
                    area_total, area_fl, area_bf, area_rentable, area_usable,
                    area_bl, area_garden, bl_to_land_ratio, fl_space_index,
                    base_rate, parcel, comments, have_type, branch_type,
                    close_type, parent_code, is_planed
                FROM bl 
                WHERE bl_id = :bl_id
            """)
            
            result = conn.execute(sql, {"bl_id": bl_id}).fetchone()
            
            if not result:
                return jsonify({'success': False, 'message': '건물 데이터를 찾을 수 없습니다.'})
            
            # 결과를 딕셔너리로 변환
            data = dict(result)
            
            # 날짜 필드 포맷팅
            date_fields = ['date_bl', 'date_buy', 'date_buy_land', 'date_sailed', 'date_manage_start']
            for field in date_fields:
                if data.get(field):
                    data[field] = data[field].strftime('%Y-%m-%d') if hasattr(data[field], 'strftime') else str(data[field])
            
            # 숫자 필드 처리
            numeric_fields = ['area_total', 'area_fl', 'area_bf', 'area_rentable', 'area_usable', 
                'area_bl', 'area_garden', 'parcel', 'price_book_value', 'price_pa_land',
                'count_fl', 'count_bf', 'bl_height', 'bl_depth', 'width_front_road',
                'width_back_road', 'width_side_road', 'parking_unit_inner', 
                'parking_unit_outdoor', 'el_unit', 'es_unit']

            for field in numeric_fields:
                if data.get(field) is not None:
                    # 원본 숫자값 그대로 반환 (콤마 없이)
                    data[field] = str(float(data[field])) if data[field] != 0 else ""

            # 소수점 필드 처리
            decimal_fields = ['bl_to_land_ratio', 'fl_space_index', 'base_rate']
            for field in decimal_fields:
                if data.get(field) is not None:
                    data[field] = str(float(data[field])) if data[field] != 0 else ""
            
            result_data = {
                'success': True,
                'data': data,
                'has_permission': has_permission
            }
            
            response = make_response(json.dumps(result_data, ensure_ascii=False))
            response.mimetype = 'application/json; charset=utf-8'
            return response
            
    except Exception as e:
        print(f"🔴 건물 데이터 조회 중 오류 발생: {str(e)}")
        return jsonify({'success': False, 'message': '서버 오류가 발생했습니다.'})

##### 건물 데이터 저장 #####
@bl_update_bp.route('/bl_update/save_bl_data', methods=['POST'])
@login_required  # 추가
def save_bl_data():
    request_data = request.get_json()
    bl_id = request_data.get('bl_id')
    em_id = request_data.get('em_id')
    
    print(f"🔵 건물 데이터 저장 요청: bl_id={bl_id}, em_id={em_id}")
    
    if not bl_id or not em_id:
        return jsonify({'success': False, 'message': 'bl_id와 em_id가 필요합니다.'})
    
    # 권한 체크
    if not check_bl_permission(em_id, bl_id):
        return jsonify({'success': False, 'message': '해당 건물의 수정 권한이 없습니다.'})
    
    try:
        with engine.connect() as conn:
            # 트랜잭션 시작
            trans = conn.begin()
            
            try:
                # 업데이트 쿼리 구성
                update_sql = text("""
                    UPDATE bl SET
                        name = :name,
                        landlord = :landlord,
                        design = :design,
                        builder = :builder,
                        address1 = :address1,
                        area_total = :area_total,
                        class_land = :class_land,
                        district = :district,
                        section = :section,
                        area_bl = :area_bl,
                        parcel = :parcel,
                        bl_to_land_ratio = :bl_to_land_ratio,
                        fl_space_index = :fl_space_index,
                        area_usable = :area_usable,
                        construction_type = :construction_type,
                        roof_type = :roof_type,
                        count_fl = :count_fl,
                        count_bf = :count_bf,
                        bl_height = :bl_height,
                        bl_depth = :bl_depth,
                        use1 = :use1,
                        date_bl = :date_bl,
                        parking_type = :parking_type,
                        elev_type = :elev_type,
                        elec_type = :elec_type,
                        generator_type = :generator_type,
                        heating_type = :heating_type,
                        cooling_type = :cooling_type,
                        contact_name = :contact_name,
                        contact_phone = :contact_phone,
                        contact_fax = :contact_fax,
                        zip = :zip,
                        regist = :regist,
                        date_buy = :date_buy,
                        date_buy_land = :date_buy_land,
                        date_sailed = :date_sailed,
                        date_manage_start = :date_manage_start,
                        price_pa_land = :price_pa_land,
                        area_fl = :area_fl,
                        area_bf = :area_bf,
                        area_rentable = :area_rentable,
                        area_garden = :area_garden,
                        price_book_value = :price_book_value,
                        base_rate = :base_rate,
                        use_fl_4 = :use_fl_4,
                        use_fl_13 = :use_fl_13,
                        width_front_road = :width_front_road,
                        width_back_road = :width_back_road,
                        width_side_road = :width_side_road,
                        parking_unit_inner = :parking_unit_inner,
                        ph = :ph,
                        parking_unit_outdoor = :parking_unit_outdoor,
                        el_unit = :el_unit,
                        es_unit = :es_unit,
                        comments = :comments,
                        em_modi = :em_modi,
                        date_modi = NOW()
                    WHERE bl_id = :bl_id
                """)
                
                # 파라미터 준비 - 숫자 필드 변환
                params = dict(request_data)
                params['bl_id'] = bl_id
                params['em_modi'] = em_id
                
                # 숫자 필드 처리 (콤마 제거)
                numeric_fields = ['area_total', 'area_fl', 'area_bf', 'area_rentable', 'area_usable', 
                                'area_bl', 'area_garden', 'parcel', 'price_book_value', 'price_pa_land',
                                'count_fl', 'count_bf', 'bl_height', 'bl_depth', 'width_front_road',
                                'width_back_road', 'width_side_road', 'parking_unit_inner', 
                                'parking_unit_outdoor', 'el_unit', 'es_unit', 'bl_to_land_ratio',
                                'fl_space_index', 'base_rate']
                
                for field in numeric_fields:
                    if params.get(field):
                        # 콤마 제거 후 숫자로 변환
                        value = str(params[field]).replace(',', '').strip()
                        params[field] = float(value) if value else None
                    else:
                        params[field] = None
                
                # 날짜 필드 처리
                date_fields = ['date_bl', 'date_buy', 'date_buy_land', 'date_sailed', 'date_manage_start']
                for field in date_fields:
                    if not params.get(field):
                        params[field] = None
                
                # 기본값 설정
                if not params.get('regist'):
                    params['regist'] = '0'
                if not params.get('use_fl_4'):
                    params['use_fl_4'] = '0'
                if not params.get('use_fl_13'):
                    params['use_fl_13'] = '0'
                if not params.get('ph'):
                    params['ph'] = '0'
                
                print(f"🟢 업데이트 파라미터: {params}")
                
                # 업데이트 실행
                result = conn.execute(update_sql, params)
                
                if result.rowcount > 0:
                    trans.commit()
                    print(f"🟢 건물 데이터 저장 완료: bl_id={bl_id}")
                    return jsonify({'success': True, 'message': '저장이 완료되었습니다.'})
                else:
                    trans.rollback()
                    return jsonify({'success': False, 'message': '저장할 데이터가 없습니다.'})
                    
            except Exception as e:
                trans.rollback()
                raise e
                
    except Exception as e:
        print(f"🔴 건물 데이터 저장 중 오류 발생: {str(e)}")
        return jsonify({'success': False, 'message': f'저장 중 오류가 발생했습니다: {str(e)}'})

##### 이력 데이터 조회 #####
@bl_update_bp.route('/bl_update/get_history_data', methods=['POST'])
@login_required  # 추가
def get_history_data():
    request_data = request.get_json()
    bl_id = request_data.get('bl_id')
    history_type = request_data.get('type', '2')  # 기본값: 이력관리
    date_start = request_data.get('date_start')
    date_end = request_data.get('date_end')
    keyword = request_data.get('keyword')
    
    print(f"🔵 이력 데이터 조회: bl_id={bl_id}, type={history_type}")
    
    if not bl_id:
        return jsonify({'success': False, 'message': 'bl_id가 필요합니다.'})
    
    try:
        with engine.connect() as conn:
            # 기본 쿼리
            sql = """
                SELECT 
                    p.auto_number,
                    p.title,
                    p.reg_man,
                    DATE(p.reg_date) as reg_date,
                    p.filename,
                    p.maskname,
                    e.name as reg_man_name
                FROM blpds p
                LEFT JOIN em e ON p.reg_man = e.em_id
                WHERE p.bl_id = :bl_id
            """
            
            params = {'bl_id': bl_id}
            
            # 파일 타입 필터
            # 파일 타입 필터
            if history_type == '1':  # 사진
                sql += " AND p.filetype = '1'"
            elif history_type == '2':  # 이력관리
                sql += " AND p.filetype = '2'"
            elif history_type == '3':  # 파일관리 (기존)
                sql += " AND p.filetype = '3'"
            elif history_type == 'file_all':  # 파일관리 (이미지+파일)
                sql += " AND (p.filetype = '1' OR p.filetype = '3')"
            
            # 날짜 필터
            if date_start:
                sql += " AND DATE(p.reg_date) >= :date_start"
                params['date_start'] = date_start
            
            if date_end:
                sql += " AND DATE(p.reg_date) <= :date_end"
                params['date_end'] = date_end
            
            # 키워드 검색
            if keyword and keyword.strip():
                sql += " AND (LOWER(p.title) LIKE LOWER(:keyword) OR LOWER(e.name) LIKE LOWER(:keyword) OR LOWER(p.contents) LIKE LOWER(:keyword))"
                params['keyword'] = f"%{keyword.strip()}%"
            
            sql += " ORDER BY p.auto_number DESC"
            
            print(f"🟢 실행할 쿼리: {sql}")
            print(f"🟢 쿼리 파라미터: {params}")
            
            result = conn.execute(text(sql), params).fetchall()
            
            # 결과 변환
            data = []
            for row in result:
                row_dict = dict(row)
                # 날짜 포맷팅
                if row_dict.get('reg_date'):
                    row_dict['reg_date'] = row_dict['reg_date'].strftime('%Y-%m-%d') if hasattr(row_dict['reg_date'], 'strftime') else str(row_dict['reg_date'])
                data.append(row_dict)
            
            result_data = {
                'success': True,
                'data': data
            }
            
            response = make_response(json.dumps(result_data, ensure_ascii=False))
            response.mimetype = 'application/json; charset=utf-8'
            return response
            
    except Exception as e:
        print(f"🔴 이력 데이터 조회 중 오류 발생: {str(e)}")
        return jsonify({'success': False, 'message': '서버 오류가 발생했습니다.'})

##### 사진 데이터 조회 #####
@bl_update_bp.route('/bl_update/get_photo_data', methods=['POST'])
@login_required  # 추가
def get_photo_data():
    request_data = request.get_json()
    bl_id = request_data.get('bl_id')
    date_start = request_data.get('date_start')
    date_end = request_data.get('date_end')
    keyword = request_data.get('keyword')
    
    print(f"🔵 사진 데이터 조회: bl_id={bl_id}")
    
    if not bl_id:
        return jsonify({'success': False, 'message': 'bl_id가 필요합니다.'})
    
    try:
        with engine.connect() as conn:
            # 메인 사진 정보 조회
            main_photo_sql = text("SELECT maskname FROM bl WHERE bl_id = :bl_id")
            main_photo_result = conn.execute(main_photo_sql, {"bl_id": bl_id}).fetchone()
            main_photo_maskname = main_photo_result['maskname'] if main_photo_result else None
            
            # 사진 목록 조회
            sql = """
                SELECT 
                    p.auto_number,
                    p.title,
                    p.reg_man,
                    DATE(p.reg_date) as reg_date,
                    p.filename,
                    p.maskname,
                    e.name as reg_man_name,
                    CASE WHEN p.maskname = :main_maskname THEN 1 ELSE 0 END as is_main_photo
                FROM blpds p
                LEFT JOIN em e ON p.reg_man = e.em_id
                WHERE p.bl_id = :bl_id 
                AND p.filetype = '1'
                AND p.maskname IS NOT NULL
            """
            
            params = {'bl_id': bl_id, 'main_maskname': main_photo_maskname}
            
            # 날짜 필터
            if date_start:
                sql += " AND DATE(p.reg_date) >= :date_start"
                params['date_start'] = date_start
            
            if date_end:
                sql += " AND DATE(p.reg_date) <= :date_end"
                params['date_end'] = date_end
            
            # 키워드 검색
            if keyword and keyword.strip():
                sql += " AND (LOWER(p.title) LIKE LOWER(:keyword) OR LOWER(e.name) LIKE LOWER(:keyword))"
                params['keyword'] = f"%{keyword.strip()}%"
            
            sql += " ORDER BY is_main_photo DESC, p.auto_number DESC"
            
            result = conn.execute(text(sql), params).fetchall()
            
            # 결과 변환
            data = []
            for row in result:
                row_dict = dict(row)
                # 날짜 포맷팅
                if row_dict.get('reg_date'):
                    row_dict['reg_date'] = row_dict['reg_date'].strftime('%Y-%m-%d') if hasattr(row_dict['reg_date'], 'strftime') else str(row_dict['reg_date'])
                data.append(row_dict)
            
            result_data = {
                'success': True,
                'data': data
            }
            
            response = make_response(json.dumps(result_data, ensure_ascii=False))
            response.mimetype = 'application/json; charset=utf-8'
            return response
            
    except Exception as e:
        print(f"🔴 사진 데이터 조회 중 오류 발생: {str(e)}")
        return jsonify({'success': False, 'message': '서버 오류가 발생했습니다.'})

##### 메인사진 삭제 #####
@bl_update_bp.route('/bl_update/delete_main_photo', methods=['POST'])
@login_required  # 추가
def delete_main_photo():
    request_data = request.get_json()
    bl_id = request_data.get('bl_id')
    
    print(f"🔵 메인사진 삭제 요청: bl_id={bl_id}")
    
    if not bl_id:
        return jsonify({'success': False, 'message': 'bl_id가 필요합니다.'})
    
    try:
        with engine.connect() as conn:
            # 트랜잭션 시작
            trans = conn.begin()
            
            try:
                # 현재 메인사진 정보 조회
                main_photo_sql = text("SELECT maskname FROM bl WHERE bl_id = :bl_id")
                main_photo_result = conn.execute(main_photo_sql, {"bl_id": bl_id}).fetchone()
                
                if not main_photo_result or not main_photo_result['maskname']:
                    trans.rollback()
                    return jsonify({'success': False, 'message': '삭제할 메인사진이 없습니다.'})
                
                maskname = main_photo_result['maskname']
                
                # bl 테이블에서 메인사진 정보 제거
                update_bl_sql = text("UPDATE bl SET maskname = NULL WHERE bl_id = :bl_id")
                conn.execute(update_bl_sql, {"bl_id": bl_id})
                
                # blpds 테이블에서 해당 사진 정보 삭제
                delete_blpds_sql = text("DELETE FROM blpds WHERE bl_id = :bl_id AND maskname = :maskname")
                conn.execute(delete_blpds_sql, {"bl_id": bl_id, "maskname": maskname})
                
                # 실제 파일 삭제 (경로는 환경에 맞게 수정)
                file_path = os.path.join('static', 'uploads', maskname)
                if os.path.exists(file_path):
                    os.remove(file_path)
                    print(f"🟢 파일 삭제 완료: {file_path}")
                
                trans.commit()
                
                return jsonify({'success': True, 'message': '메인사진이 삭제되었습니다.'})
                
            except Exception as e:
                trans.rollback()
                raise e
                
    except Exception as e:
        print(f"🔴 메인사진 삭제 중 오류 발생: {str(e)}")
        return jsonify({'success': False, 'message': f'메인사진 삭제 중 오류가 발생했습니다: {str(e)}'})

##### 엑셀 내보내기 #####
@bl_update_bp.route('/bl_update/export_excel', methods=['GET'])
@login_required  # 추가
def export_excel():
    bl_id = request.args.get('bl_id')
    
    if not bl_id:
        return jsonify({'success': False, 'message': 'bl_id가 필요합니다.'})
    
    try:
        with engine.connect() as conn:
            sql = text("""
                SELECT * FROM bl WHERE bl_id = :bl_id
            """)
            
            result = conn.execute(sql, {"bl_id": bl_id}).fetchone()
            
            if not result:
                return jsonify({'success': False, 'message': '건물 데이터를 찾을 수 없습니다.'})
            
            # 엑셀 파일 생성
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "건물정보"
            
            # 헤더 설정
            headers = [
                '건물ID', '건물명', '소재지', '연면적', '건축면적', '대지면적',
                '지상층수', '지하층수', '준공일', '건축구조', '소유주',
                '설계자', '시공사', '담당자', '연락처'
            ]
            
            # 헤더 추가
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 데이터 추가
            data = dict(result)
            row_data = [
                data.get('bl_id', ''),
                data.get('name', ''),
                data.get('address1', ''),
                data.get('area_total', ''),
                data.get('area_bl', ''),
                data.get('parcel', ''),
                data.get('count_fl', ''),
                data.get('count_bf', ''),
                data.get('date_bl', ''),
                data.get('construction_type', ''),
                data.get('landlord', ''),
                data.get('design', ''),
                data.get('builder', ''),
                data.get('contact_name', ''),
                data.get('contact_phone', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=2, column=col, value=value)
            
            # 메모리에 파일 저장
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 파일명 생성
            filename = f"building_info_{bl_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
    except Exception as e:
        print(f"🔴 엑셀 내보내기 중 오류 발생: {str(e)}")
        return jsonify({'success': False, 'message': '엑셀 내보내기 중 오류가 발생했습니다.'})
def safe_int(value):
    """값을 안전하게 정수로 변환"""
    try:
        if value is None:
            return None
        # float이든 string이든 일단 float으로 변환 후 int로 변환
        return int(float(value))
    except (ValueError, TypeError):
        return None
##### 파일 보기 #####
BL_PDS_PATH = r"C:\Users\USER04\Documents\python_fms_hiddenframe\upload\bl_pds"

@bl_update_bp.route('/bl_update/view_file', methods=['GET'])
@login_required
def view_file():
    auto_number_param = request.args.get('auto_number')
    auto_number = safe_int(auto_number_param)
    
    print(f"🏢 [bl_update] view_file 폴백 엔드포인트 호출: auto_number={auto_number}")
    
    if not auto_number:
        print("🔴 auto_number가 없음")
        return send_file('static/images/common/no_image.png')
    
    try:
        with engine.connect() as conn:
            sql = text("""
                SELECT filename, maskname, title, filetype
                FROM blpds 
                WHERE auto_number = :auto_number
            """)
            
            result = conn.execute(sql, {"auto_number": auto_number}).fetchone()
            
            if not result or not result['maskname']:
                print(f"🔴 파일 정보 없음: auto_number={auto_number}")
                return send_file('static/images/common/no_image.png')
            
            file_data = dict(result)
            maskname = file_data['maskname']
            
            # ⭐ bl_pds 절대 경로에서 파일 찾기
            bl_pds_file_path = os.path.join(BL_PDS_PATH, maskname)
            
            print(f"🏢 [bl_update] 폴백에서 파일 경로 확인: {bl_pds_file_path}")
            
            if os.path.exists(bl_pds_file_path):
                print(f"✅ 파일 찾음: {bl_pds_file_path}")
                
                # 파일 확장자에 따른 mimetype 설정
                file_ext = maskname.lower().split('.')[-1]
                mimetype_map = {
                    'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png',
                    'gif': 'image/gif', 'bmp': 'image/bmp', 'webp': 'image/webp',
                    'svg': 'image/svg+xml'
                }
                
                mimetype = mimetype_map.get(file_ext, 'application/octet-stream')
                
                return send_file(
                    bl_pds_file_path,
                    mimetype=mimetype,
                    as_attachment=False,
                    download_name=file_data['filename']
                )
            else:
                print(f"🔴 파일 없음: {bl_pds_file_path}")
                return send_file('static/images/common/no_image.png')
            
    except Exception as e:
        print(f"🔴 [bl_update] 폴백 엔드포인트 오류: {str(e)}")
        return send_file('static/images/common/no_image.png')


@bl_update_bp.route('/bl_update/debug_file/<int:auto_number>')    
@login_required
def debug_file(auto_number):
    """파일 존재 여부 디버깅"""
    try:
        with engine.connect() as conn:
            sql = text("""
                SELECT auto_number, filename, maskname, filetype, title
                FROM blpds 
                WHERE auto_number = :auto_number
            """)
            
            result = conn.execute(sql, {"auto_number": auto_number}).fetchone()
            
            debug_info = {
                'auto_number': auto_number,
                'database_record': dict(result) if result else None,
                'current_directory': os.getcwd(),
                'file_checks': [],
                'uploads_directory_exists': False,
                'uploads_files': []
            }
            
            if result:
                maskname = result['maskname']
                if maskname:
                    # 여러 경로 확인
                    possible_paths = [
                        os.path.join('static', 'uploads', maskname),
                        os.path.join('uploads', maskname),
                        maskname,
                        os.path.join(os.getcwd(), 'static', 'uploads', maskname)
                    ]
                    
                    for path in possible_paths:
                        exists = os.path.exists(path)
                        size = os.path.getsize(path) if exists else 0
                        debug_info['file_checks'].append({
                            'path': path,
                            'exists': exists,
                            'size': size
                        })
            
            # uploads 디렉토리 확인
            uploads_dir = os.path.join('static', 'uploads')
            if os.path.exists(uploads_dir):
                debug_info['uploads_directory_exists'] = True
                try:
                    files = os.listdir(uploads_dir)[:20]  # 최대 20개
                    debug_info['uploads_files'] = files
                except:
                    debug_info['uploads_files'] = ['읽기 오류']
            
            return jsonify(debug_info)
            
    except Exception as e:
        return jsonify({
            'error': str(e),
            'auto_number': auto_number
        })
        
@bl_update_bp.route('/bl_update/export_excel_detailed', methods=['POST'])
@login_required
def export_excel_detailed():
    request_data = request.get_json()
    bl_id = request_data.get('bl_id')
    include_history = request_data.get('include_history', False)
    
    print(f"🔵 상세 엑셀 내보내기: bl_id={bl_id}, include_history={include_history}")
    
    if not bl_id:
        return jsonify({'success': False, 'message': 'bl_id가 필요합니다.'})
    
    try:
        with engine.connect() as conn:
            # 건물 기본 정보 조회
            bl_sql = text("""
                SELECT 
                    bl_id, prop_id, name, zip, contact_phone, address1, address2,
                    use1, contact_fax, DATE(date_bl) as date_bl, price_book_value,
                    contact_name, maskname, count_fl, count_bf, construction_type,
                    use_fl_4, use_fl_13, bl_height, bl_depth, width_front_road,
                    width_back_road, width_side_road, ph, el, el_unit, parking_type,
                    es, es_unit, parking_unit_inner, parking_unit_outdoor,
                    cooling_type, heating_type, DATE(date_buy) as date_buy,
                    DATE(date_buy_land) as date_buy_land, DATE(date_sailed) as date_sailed,
                    DATE(date_manage_start) as date_manage_start, price_pa_land,
                    class_land, regist, district, section, landlord, design,
                    builder, elev_type, elec_type, generator_type, roof_type,
                    area_total, area_fl, area_bf, area_rentable, area_usable,
                    area_bl, area_garden, bl_to_land_ratio, fl_space_index,
                    base_rate, parcel, comments, have_type, branch_type,
                    close_type, parent_code, is_planed
                FROM bl 
                WHERE bl_id = :bl_id
            """)
            
            bl_result = conn.execute(bl_sql, {"bl_id": bl_id}).fetchone()
            
            if not bl_result:
                return jsonify({'success': False, 'message': '건물 데이터를 찾을 수 없습니다.'})
            
            bl_data = dict(bl_result)
            
            # 엑셀 파일 생성
            wb = openpyxl.Workbook()
            
            # 첫 번째 시트: 건물 개요
            ws1 = wb.active
            ws1.title = "건물개요"
            
            # 건물 개요 헤더
            ws1.merge_cells('A1:D1')
            ws1['A1'] = f"건물 정보 상세 - {bl_data.get('name', '')}"
            ws1['A1'].font = openpyxl.styles.Font(size=16, bold=True)
            ws1['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # 기본 정보
            basic_info = [
                ('건물명', bl_data.get('name', '')),
                ('건물ID', bl_data.get('bl_id', '')),
                ('소재지', bl_data.get('address1', '')),
                ('우편번호', bl_data.get('zip', '')),
                ('소유주', bl_data.get('landlord', '')),
                ('설계자', bl_data.get('design', '')),
                ('시공사', bl_data.get('builder', '')),
                ('준공일', bl_data.get('date_bl', '')),
                ('주용도', bl_data.get('use1', '')),
                ('담당자', bl_data.get('contact_name', '')),
                ('연락처', bl_data.get('contact_phone', '')),
                ('FAX', bl_data.get('contact_fax', ''))
            ]
            
            row = 3
            for label, value in basic_info:
                ws1[f'A{row}'] = label
                ws1[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                ws1[f'A{row}'].fill = openpyxl.styles.PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
                ws1[f'B{row}'] = value
                row += 1
            
            # 두 번째 시트: 면적 및 구조 정보
            ws2 = wb.create_sheet("면적및구조")
            
            ws2.merge_cells('A1:D1')
            ws2['A1'] = "면적 및 구조 정보"
            ws2['A1'].font = openpyxl.styles.Font(size=14, bold=True)
            ws2['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
            
            area_info = [
                ('연면적(㎡)', bl_data.get('area_total', '')),
                ('건축면적(㎡)', bl_data.get('area_bl', '')),
                ('대지면적(㎡)', bl_data.get('parcel', '')),
                ('지상층면적(㎡)', bl_data.get('area_fl', '')),
                ('지하층면적(㎡)', bl_data.get('area_bf', '')),
                ('임대가능면적(㎡)', bl_data.get('area_rentable', '')),
                ('전용면적(㎡)', bl_data.get('area_usable', '')),
                ('조경면적(㎡)', bl_data.get('area_garden', '')),
                ('지상층수', bl_data.get('count_fl', '')),
                ('지하층수', bl_data.get('count_bf', '')),
                ('최고높이(m)', bl_data.get('bl_height', '')),
                ('지하깊이(m)', bl_data.get('bl_depth', '')),
                ('건축구조', bl_data.get('construction_type', '')),
                ('지붕형태', bl_data.get('roof_type', '')),
                ('건폐율(%)', bl_data.get('bl_to_land_ratio', '')),
                ('용적율(%)', bl_data.get('fl_space_index', ''))
            ]
            
            row = 3
            for label, value in area_info:
                ws2[f'A{row}'] = label
                ws2[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                ws2[f'A{row}'].fill = openpyxl.styles.PatternFill(start_color='F0F8E8', end_color='F0F8E8', fill_type='solid')
                ws2[f'B{row}'] = value
                row += 1
            
            # 세 번째 시트: 설비 및 기타 정보
            ws3 = wb.create_sheet("설비및기타")
            
            ws3.merge_cells('A1:D1')
            ws3['A1'] = "설비 및 기타 정보"
            ws3['A1'].font = openpyxl.styles.Font(size=14, bold=True)
            ws3['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
            
            facility_info = [
                ('주차설비', bl_data.get('parking_type', '')),
                ('승강설비', bl_data.get('elev_type', '')),
                ('수전설비', bl_data.get('elec_type', '')),
                ('발전설비', bl_data.get('generator_type', '')),
                ('난방설비', bl_data.get('heating_type', '')),
                ('냉방설비', bl_data.get('cooling_type', '')),
                ('옥내주차대수', bl_data.get('parking_unit_inner', '')),
                ('옥외주차대수', bl_data.get('parking_unit_outdoor', '')),
                ('E/L대수', bl_data.get('el_unit', '')),
                ('E/S대수', bl_data.get('es_unit', '')),
                ('전면도로폭(m)', bl_data.get('width_front_road', '')),
                ('후면도로폭(m)', bl_data.get('width_back_road', '')),
                ('측면도로폭(m)', bl_data.get('width_side_road', '')),
                ('지목', bl_data.get('class_land', '')),
                ('지역', bl_data.get('district', '')),
                ('지구', bl_data.get('section', '')),
                ('건물취득일', bl_data.get('date_buy', '')),
                ('대지취득일', bl_data.get('date_buy_land', '')),
                ('관리개시일', bl_data.get('date_manage_start', '')),
                ('매각일자', bl_data.get('date_sailed', '')),
                ('장부가(원)', bl_data.get('price_book_value', '')),
                ('공시지가(원)', bl_data.get('price_pa_land', '')),
                ('비고', bl_data.get('comments', ''))
            ]
            
            row = 3
            for label, value in facility_info:
                ws3[f'A{row}'] = label
                ws3[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                ws3[f'A{row}'].fill = openpyxl.styles.PatternFill(start_color='F8F0E8', end_color='F8F0E8', fill_type='solid')
                ws3[f'B{row}'] = value
                row += 1
            
            # 이력 정보 포함
            if include_history:
                # 이력 데이터 조회
                history_sql = text("""
                    SELECT 
                        p.auto_number,
                        p.title,
                        p.contents,
                        p.reg_man,
                        DATE(p.reg_date) as reg_date,
                        p.filename,
                        p.filetype,
                        e.name as reg_man_name,
                        CASE 
                            WHEN p.filetype = '1' THEN '이미지'
                            WHEN p.filetype = '2' THEN '텍스트'
                            WHEN p.filetype = '3' THEN '파일'
                            ELSE '기타'
                        END as filetype_name
                    FROM blpds p
                    LEFT JOIN em e ON p.reg_man = e.em_id
                    WHERE p.bl_id = :bl_id
                    ORDER BY p.reg_date DESC
                """)
                
                history_result = conn.execute(history_sql, {"bl_id": bl_id}).fetchall()
                
                if history_result:
                    ws4 = wb.create_sheet("건물이력")
                    
                    # 헤더
                    headers = ['번호', '제목', '내용', '유형', '등록자', '등록일', '파일명']
                    for col, header in enumerate(headers, 1):
                        cell = ws4.cell(row=1, column=col, value=header)
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(start_color='E8E8F8', end_color='E8E8F8', fill_type='solid')
                    
                    # 데이터
                    for row_num, row_data in enumerate(history_result, 2):
                        ws4.cell(row=row_num, column=1, value=row_data['auto_number'])
                        ws4.cell(row=row_num, column=2, value=row_data['title'] or '')
                        ws4.cell(row=row_num, column=3, value=row_data['contents'] or '')
                        ws4.cell(row=row_num, column=4, value=row_data['filetype_name'])
                        ws4.cell(row=row_num, column=5, value=row_data['reg_man_name'] or row_data['reg_man'])
                        ws4.cell(row=row_num, column=6, value=row_data['reg_date'].strftime('%Y-%m-%d') if row_data['reg_date'] else '')
                        ws4.cell(row=row_num, column=7, value=row_data['filename'] or '')
                    
                    # 열 너비 조정
                    ws4.column_dimensions['B'].width = 30  # 제목
                    ws4.column_dimensions['C'].width = 50  # 내용
                    ws4.column_dimensions['G'].width = 25  # 파일명
            
            # 모든 시트의 열 너비 조정
            for ws in [ws1, ws2, ws3]:
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 30
            
            # 메모리에 파일 저장
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 파일명 생성
            bl_name = bl_data.get('name', '건물정보').replace('/', '_').replace('\\', '_')
            filename = f"건물정보상세_{bl_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            print(f"🟢 상세 엑셀 생성 완료: {filename}")
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
    except Exception as e:
        print(f"🔴 상세 엑셀 내보내기 중 오류 발생: {str(e)}")
        return jsonify({'success': False, 'message': f'엑셀 내보내기 중 오류가 발생했습니다: {str(e)}'})